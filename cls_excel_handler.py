from openpyxl import Workbook
from openpyxl.compat import range
import os
import tempfile
from typing import Union
from typing import List
from openpyxl import load_workbook


class ClsExcelHandler():
    def __init__(self):
        self._books_path = tempfile.gettempdir()

    def __add_book(self):
        """新增活頁簿"""
        self.__book = Workbook()
        self.__sheet = self.__book.active

    def save_book(self, book_path: str):
        """儲存活頁簿

            Arguments:
                book_path {str} -- 本機路徑
        """
        self.__book.save(book_path)

    def write_to_sheet(self, values: Union[List[List[str]], List[str], str]):
        """寫入工作表

            Arguments:
                values {Union[List[List[str]], List[str], str]} -- 要寫入的值
        """
        if type(values) is List[List[str]]:
            for currentIndex in range(1, len(values)):
                self.__sheet.append(values[currentIndex])
        elif type(values) is List[str] or type(values) is str:
            self.__sheet.append(values)
        else:
            raise ValueError('values型別只能是(List[List[str]]/List[str]/str)其中之一')

    def open_books_directory(self, books_path: str):
        """開啟活頁簿預設儲存目錄

            Arguments:
                books_path {str} -- 本機路徑
        """
        if not os.path.exists(books_path):
            os.makedirs(books_path)
        self._books_path = books_path

    def open_book(self, book_path: str):
        """開啟活頁簿(不存在則先建立)

            Arguments:
                book_path {str} -- 本機路徑
        """
        if not self.is_book_existed(book_path):
            self.__add_book()
        else:
            self.__book = load_workbook(book_path)
            self.__sheet = self.__book.active

    def __add_sheet(self, sheet_name: str):
        """新增工作表

            Arguments:
                sheet_name {str} -- 工作表名稱
        """
        self.__book.create_sheet(sheet_name)

    def open_sheet(self, sheet_name: str):
        """開啟工作表(不存在則先建立)

            Arguments:
                sheet_name {str} -- 工作表名稱
        """
        if not self.is_sheet_existed():
            self.__add_sheet(sheet_name)
        self.__book.active = self.__book.worksheets.index(
            self.get_sheet_by_name(sheet_name))

    def is_book_existed(self, book_path: str) -> bool:
        """判斷活頁簿是否存在

            Arguments:
                book_path {str} -- 本機路徑

            Returns:
                bool -- 回傳結果
        """
        return os.path.exists(book_path)

    def is_sheet_existed(self, sheet_name: str) -> bool:
        """判斷工作表是否存在

            Arguments:
                sheet_name {str} -- 工作表名稱

            Returns:
                bool -- 回傳結果
        """
        if sheet_name in self.__book.sheetnames:
            return True
        else:
            return False
